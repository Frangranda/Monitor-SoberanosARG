"""
Script para actualizar historicos.xlsx con precios de cierre del día.
Se ejecuta automáticamente via GitHub Actions al cierre del mercado.
Los tickers se leen dinámicamente desde Instrumentos.xlsx.
"""

import requests
import openpyxl
from openpyxl import load_workbook
from datetime import date
import os
import re
import time

# ── CONFIGURACIÓN ─────────────────────────────────────────────
ECO_BASE        = "https://bonos.ecovalores.com.ar/eco/ticker.php"
HISTORICOS_FILE = "historicos.xlsx"
INSTRUMENTOS_FILE = "Instrumentos.xlsx"

# Grupos que usan sufijo D para buscar el precio en USD
GRUPOS_CON_D = {'USD Bonares', 'USD Globales'}

# ── LEER TICKERS DESDE INSTRUMENTOS.XLSX ─────────────────────
def leer_tickers():
    if not os.path.exists(INSTRUMENTOS_FILE):
        print(f"ERROR: No se encontró {INSTRUMENTOS_FILE}")
        return []

    wb = load_workbook(INSTRUMENTOS_FILE, read_only=True, data_only=True)
    tickers = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Buscar fila de header (contiene 'Ticker')
        header_row = None
        for i, row in enumerate(rows):
            if row and str(row[0]).strip() == 'Ticker':
                header_row = i
                break

        if header_row is None:
            continue

        headers = [str(c).strip() if c else '' for c in rows[header_row]]
        ticker_col = headers.index('Ticker') if 'Ticker' in headers else 0

        for row in rows[header_row + 1:]:
            if not row or not row[ticker_col]:
                continue
            ticker = str(row[ticker_col]).strip()
            if not ticker or ticker == 'Ticker' or ticker == 'None':
                continue

            # Agregar sufijo D para Bonares y Globales
            eco_ticker = ticker + 'D' if sheet_name in GRUPOS_CON_D else ticker

            if eco_ticker not in tickers:
                tickers.append(eco_ticker)

    print(f"Tickers leídos desde {INSTRUMENTOS_FILE}: {len(tickers)}")
    return tickers

# ── FETCH PRECIO ──────────────────────────────────────────────
def fetch_precio(ticker):
    try:
        url = f"{ECO_BASE}?t={ticker}"
        resp = requests.get(url, timeout=10, headers={
            "User-Agent": "Mozilla/5.0",
            "Accept": "text/html",
            "Referer": "https://bonos.ecovalores.com.ar"
        })
        html = resp.text
        match = re.search(r'<td class="precioticker">\s*([\d.,]+)\s*</td>', html)
        if match:
            price_str = match.group(1).replace(".", "").replace(",", ".")
            return float(price_str)
    except Exception as e:
        print(f"  Error fetching {ticker}: {e}")
    return None

# ── ACTUALIZAR EXCEL ──────────────────────────────────────────
def actualizar_historicos():
    hoy = date.today()
    fecha_str = hoy.strftime("%Y-%m-%d")
    print(f"Actualizando historicos para {fecha_str}...")

    # Leer tickers dinámicamente
    tickers = leer_tickers()
    if not tickers:
        print("ERROR: No se pudieron leer los tickers.")
        return

    # Cargar o crear el Excel
    if os.path.exists(HISTORICOS_FILE):
        wb = load_workbook(HISTORICOS_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historicos"
        ws.cell(row=1, column=1, value="Fecha")
        print("Archivo historicos.xlsx creado desde cero.")

    # Verificar si ya existe la fila de hoy
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0])[:10] == fecha_str:
            print(f"Ya existe fila para {fecha_str}, saliendo.")
            return

    # Próxima fila vacía
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=fecha_str)

    # Asegurar que todos los tickers estén en el header
    header = {ws.cell(row=1, column=c).value: c for c in range(2, ws.max_column + 1)}
    for ticker in tickers:
        if ticker not in header:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=ticker)
            header[ticker] = new_col
            print(f"  Nuevo ticker agregado al header: {ticker}")

    # Fetchear precios
    ok = 0
    err = 0
    for ticker in tickers:
        print(f"  Fetching {ticker}...", end=" ")
        precio = fetch_precio(ticker)
        if precio:
            col = header[ticker]
            ws.cell(row=next_row, column=col, value=precio)
            print(f"${precio}")
            ok += 1
        else:
            print("sin precio")
            err += 1
        time.sleep(0.4)

    wb.save(HISTORICOS_FILE)
    print(f"\nListo: {ok} precios guardados, {err} sin precio.")

if __name__ == "__main__":
    actualizar_historicos()
